#!/usr/bin/env python3
"""
update_tracker.py
Regenerates data/grocery_tracker.xlsx from the markdown source-of-truth files
in data/grocery_trips/, data/meal_logs/, data/waste_logs/,
data/inventory_snapshot.md, and data/reconciliation_log.md.

Run manually:  python scripts/update_tracker.py
Run in CI:     triggered by .github/workflows/update-tracker.yml
"""

import os
import re
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
OUT_PATH = os.path.join(DATA, "grocery_tracker.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def make_sheet(wb, name, headers, col_widths=None):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def parse_markdown_table(text):
    """Extract rows from the first markdown table found in a text block."""
    lines = [l for l in text.splitlines() if l.strip().startswith("|")]
    rows = []
    for l in lines:
        cells = [c.strip() for c in l.strip().strip("|").split("|")]
        if all(re.fullmatch(r"-+", c) for c in cells if c):
            continue  # skip separator row
        rows.append(cells)
    return rows


def parse_purchases():
    """Parse data/grocery_trips/*.md -> list of purchase rows."""
    rows = []
    for path in sorted(glob.glob(os.path.join(DATA, "grocery_trips", "*.md"))):
        date = os.path.splitext(os.path.basename(path))[0]
        with open(path, encoding="utf-8") as f:
            text = f.read()
        store_match = re.search(r"\*\*Store:\*\*\s*(.+)", text)
        trip_match = re.search(r"\*\*Trip ID:\*\*\s*(\S+)", text)
        store = store_match.group(1).strip() if store_match else ""
        trip_id = trip_match.group(1).strip() if trip_match else ""
        table = parse_markdown_table(text)
        if len(table) < 2:
            continue
        header = table[0]
        for r in table[1:]:
            if len(r) < len(header):
                continue
            rowdict = dict(zip(header, r))
            item = rowdict.get("Item", "")
            if not item:
                continue
            rows.append([
                trip_id, rowdict.get("Batch ID Assigned", ""), date, store,
                item, rowdict.get("Brand/Notes", ""), rowdict.get("Qty", ""),
                rowdict.get("Unit", ""), rowdict.get("Total Price", ""),
                rowdict.get("Unit Cost", ""),
            ])
    return rows


def parse_meals():
    """Parse data/meal_logs/*.md -> (meal rows, usage rows)."""
    meal_rows, usage_rows = [], []
    meal_counter = 1
    for path in sorted(glob.glob(os.path.join(DATA, "meal_logs", "*.md"))):
        date = os.path.splitext(os.path.basename(path))[0]
        with open(path, encoding="utf-8") as f:
            text = f.read()
        sections = re.split(r"^## ", text, flags=re.MULTILINE)[1:]
        for sec in sections:
            lines = sec.splitlines()
            meal_type = lines[0].strip()
            if meal_type.lower() not in ("breakfast", "lunch", "dinner", "snacks", "snack"):
                continue
            desc_match = re.search(r"\*\*Description:\*\*\s*(.+)", sec)
            cost_match = re.search(r"\*\*Estimated Meal Cost:\*\*\s*(\$?[\d.]+)", sec)
            description = desc_match.group(1).strip() if desc_match else ""
            cost = cost_match.group(1).strip() if cost_match else ""
            meal_id = f"M{meal_counter:04d}"
            meal_counter += 1
            meal_rows.append([meal_id, date, meal_type, description, cost, "", ""])
            table = parse_markdown_table(sec)
            if len(table) >= 2:
                header = table[0]
                for r in table[1:]:
                    if len(r) < len(header):
                        continue
                    rowdict = dict(zip(header, r))
                    item = rowdict.get("Item", "")
                    if not item:
                        continue
                    usage_rows.append([
                        meal_id, rowdict.get("Batch Used", ""), item,
                        rowdict.get("Inferred Qty", ""), "", rowdict.get("Depletion Applied", ""),
                        rowdict.get("Fuzzy?", ""), rowdict.get("Notes", ""),
                    ])
    return meal_rows, usage_rows


def parse_waste():
    """Parse data/waste_logs/*.md -> list of waste rows."""
    rows = []
    for path in sorted(glob.glob(os.path.join(DATA, "waste_logs", "*.md"))):
        date = os.path.splitext(os.path.basename(path))[0]
        with open(path, encoding="utf-8") as f:
            text = f.read()
        raw_match = re.search(r'\*\*Raw Statement:\*\*\s*"(.+)"', text)
        raw = raw_match.group(1).strip() if raw_match else ""
        table = parse_markdown_table(text)
        if len(table) < 2:
            continue
        header = table[0]
        for r in table[1:]:
            if len(r) < len(header):
                continue
            rowdict = dict(zip(header, r))
            item = rowdict.get("Item", "")
            if not item:
                continue
            rows.append([
                date, item, rowdict.get("Batch ID", ""), rowdict.get("Qty Wasted", ""),
                rowdict.get("Unit", ""), rowdict.get("Unit Cost", ""),
                rowdict.get("Waste Value", ""), rowdict.get("Reason", ""), raw,
            ])
    return rows


def parse_inventory():
    """Parse data/inventory_snapshot.md -> list of inventory rows."""
    rows = []
    path = os.path.join(DATA, "inventory_snapshot.md")
    if not os.path.exists(path):
        return rows
    with open(path, encoding="utf-8") as f:
        text = f.read()
    blocks = re.split(r"^### ", text, flags=re.MULTILINE)[1:]
    for block in blocks:
        lines = block.splitlines()
        item_name = lines[0].strip()
        table = parse_markdown_table(block)
        if len(table) < 2:
            continue
        header = table[0]
        for r in table[1:]:
            if len(r) < len(header):
                continue
            rowdict = dict(zip(header, r))
            rows.append([
                rowdict.get("Batch ID", ""), item_name, rowdict.get("Date Purchased", ""),
                rowdict.get("Store", ""), rowdict.get("Qty Purchased", ""), rowdict.get("Unit", ""),
                rowdict.get("Unit Cost", ""), rowdict.get("Qty Remaining", ""),
                rowdict.get("Status", ""), "", rowdict.get("Notes", ""),
            ])
    return rows


def parse_reconciliation():
    """Parse data/reconciliation_log.md -> list of reconciliation rows."""
    rows = []
    path = os.path.join(DATA, "reconciliation_log.md")
    if not os.path.exists(path):
        return rows
    with open(path, encoding="utf-8") as f:
        text = f.read()
    blocks = re.split(r"^### ", text, flags=re.MULTILINE)[1:]
    for block in blocks:
        header_line = block.splitlines()[0].strip()
        m = re.match(r"(.+?)\s*—\s*Batch\s*(\S+)\s*—\s*(.+)", header_line)
        item, batch_id = (m.group(1), m.group(2)) if m else (header_line, "")
        status = re.search(r"\*\*Status:\*\*\s*(\S+)", block)
        total_cost = re.search(r"\*\*Batch Total Cost:\*\*\s*(\$?[\d.]+)", block)
        qty = re.search(r"\*\*Batch Qty:\*\*\s*(\S+)", block)
        reconciled_on = re.search(r"\*\*Reconciled On:\*\*\s*(\S+)", block)
        cost_split = re.search(r"\*\*Cost Split:\*\*\s*(.+)", block)
        meals_block = re.search(r"\*\*Meals Using This Batch:\*\*\s*((?:\s*-.+\n?)+)", block)
        meal_count = len(re.findall(r"^\s*-\s", meals_block.group(1), flags=re.MULTILINE)) if meals_block else 0
        rows.append([
            batch_id, item, status.group(1) if status else "",
            total_cost.group(1) if total_cost else "", qty.group(1) if qty else "",
            meal_count, cost_split.group(1) if cost_split else "",
            reconciled_on.group(1) if reconciled_on else "", "",
        ])
    return rows


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = make_sheet(wb, "Purchases",
        ["Trip_ID","Batch_ID","Date","Store","Item","Brand_Notes","Qty_Purchased","Unit","Total_Price","Unit_Cost"],
        [10,10,12,15,20,20,14,10,12,10])
    for r in parse_purchases():
        ws.append(r)

    ws = make_sheet(wb, "Inventory",
        ["Batch_ID","Item","Date_Purchased","Store","Qty_Purchased","Unit","Unit_Cost","Qty_Remaining","Status","Fuzzy_Item","Notes"],
        [10,20,14,15,14,10,10,14,12,10,25])
    for r in parse_inventory():
        ws.append(r)

    ws = make_sheet(wb, "Meals",
        ["Meal_ID","Date","Meal_Type","Description","Confirmed_Cost","Fuzzy_Pending","Notes"],
        [10,12,12,35,14,14,25])
    meal_rows, usage_rows = parse_meals()
    for r in meal_rows:
        ws.append(r)

    ws = make_sheet(wb, "Meal_Usage",
        ["Meal_ID","Batch_ID","Item","Inferred_Qty","Unit","Cost_Share","Fuzzy","Notes"],
        [10,10,20,14,10,12,10,25])
    for r in usage_rows:
        ws.append(r)

    ws = make_sheet(wb, "Waste",
        ["Date","Item","Batch_ID","Qty_Wasted","Unit","Unit_Cost","Waste_Value","Reason","Raw_Statement"],
        [12,20,10,12,10,10,12,18,30])
    for r in parse_waste():
        ws.append(r)

    ws = make_sheet(wb, "Reconciliation",
        ["Batch_ID","Item","Status","Batch_Total_Cost","Batch_Qty","Meal_Count","Cost_Per_Meal","Reconciled_Date","Notes"],
        [10,20,12,16,10,12,14,16,25])
    for r in parse_reconciliation():
        ws.append(r)

    ws = make_sheet(wb, "Dashboard", ["Stat","Value","Description"], [28,16,45])
    stats = [
        ("total_spent","=SUM(Purchases.I:I)","Cumulative grocery spend"),
        ("total_meals_logged","=COUNTA(Meals.A2:A10000)","Count of logged meals"),
        ("avg_cost_per_meal","=IFERROR(AVERAGE(Meals.E:E),0)","Rolling average cost across all meals"),
        ("avg_cost_breakfast","=IFERROR(AVERAGEIF(Meals.C:C,\"Breakfast\",Meals.E:E),0)","Avg cost for breakfast meals"),
        ("avg_cost_lunch","=IFERROR(AVERAGEIF(Meals.C:C,\"Lunch\",Meals.E:E),0)","Avg cost for lunch meals"),
        ("avg_cost_dinner","=IFERROR(AVERAGEIF(Meals.C:C,\"Dinner\",Meals.E:E),0)","Avg cost for dinner meals"),
        ("total_waste_dollars","=SUM(Waste.G:G)","Cumulative waste in dollars"),
        ("waste_pct","=IFERROR(Dashboard.B7/Dashboard.B2,0)","Waste as % of total spend"),
        ("total_trips","=SUMPRODUCT((Purchases.A2:A10000<>\"\")/COUNTIF(Purchases.A2:A10000,Purchases.A2:A10000&\"\"))","Distinct grocery trips"),
        ("avg_spend_per_trip","=IFERROR(Dashboard.B2/Dashboard.B9,0)","Average receipt total per trip"),
        ("meals_per_grocery_trip","=IFERROR(Dashboard.B3/Dashboard.B9,0)","Meals consumed per trip window"),
    ]
    for i, (stat, formula, desc) in enumerate(stats, start=2):
        ws.cell(row=i, column=1, value=stat)
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=3, value=desc)

    order = ["Purchases","Inventory","Meals","Meal_Usage","Waste","Reconciliation","Dashboard"]
    wb._sheets = [wb[s] for s in order]
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    os.makedirs(DATA, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH}")
